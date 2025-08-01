from tkinter import *
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from datetime import datetime
import mysql.connector


# ================== Database Connection ==================
def connect_db():
    try:
        conn = mysql.connector.connect(
            host="127.0.0.1",
            port=3306,
            user="root",
            password="Shyam123",
            database="champs",
            use_pure=True,
            connection_timeout=5
        )
        return conn
    except mysql.connector.Error as err:
        return None


root = Tk()
root.title("Inventory Management System")
root.geometry("1000x700")
root.configure(bg="skyblue")

# ✅ Treeview Style Configuration
style = ttk.Style()
style.theme_use("default")

style.configure("Custom.Treeview",
    background="black",
    foreground="blue",
    fieldbackground="black",
    rowheight=25)

style.configure("Custom.Treeview.Heading",
    background="gray20",
    foreground="white",
    font=('Arial', 10, 'bold'))

# ✅ Table Frame
table_frame = Frame(root, bg="white")
table_frame.place(x=20, y=360, width=960, height=280)

# ✅ Treeview Columns
columns = ("ID", "Product", "Qty", "Date", "Transaction Type", "Rate", "Amount")
tree = ttk.Treeview(table_frame, columns=columns, show="headings", style="Custom.Treeview")
tree.pack(fill=BOTH, expand=1)

# ✅ Define Headings
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, anchor=CENTER, width=120)


# ================== Add Data Function ==================
def add_transaction():

    # ---- 1) Read & validate inputs ----
    product = product_cb.get().strip()
    trans_type = trans_cb.get().strip()

    if not product:
        messagebox.showwarning("Validation", "Product select/enter karo.")
        product_cb.focus_set()
        return

    if trans_type not in ("Sale", "Purchase"):
        messagebox.showwarning("Validation", "Transaction Type select karo (Sale / Purchase).")
        trans_cb.focus_set()
        return

    try:
        qty = int(qty_entry.get())
        if qty <= 0:
            raise ValueError
    except ValueError:
        messagebox.showwarning("Validation", "Quantity valid positive integer hona chahiye.")
        qty_entry.focus_set()
        return

    try:
        rate = float(rate_entry.get())
        if rate <= 0:
            raise ValueError
    except ValueError:
        messagebox.showwarning("Validation", "Rate valid positive number hona chahiye.")
        rate_entry.focus_set()
        return

    # tkcalendar ka DateEntry safest: get_date() (python date object deta hai)
    try:
        date_obj = date_entry.get_date()  # returns datetime.date
    except Exception:
        # Fallback if you're using get()
        from datetime import datetime
        date_obj = datetime.strptime(date_entry.get(), "%m/%d/%y").date()

    amount = round(qty * rate, 2)


    # ---- 2) DB work ----
    conn = None
    cursor = None
    try:
        conn = connect_db()
        if conn is None:
            print("⛔ DB connection is None, aborting insert.")
            return

        cursor = conn.cursor()
        sql = """
            INSERT INTO transactions (product, qty, date, transaction_type, rate, amount)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(sql, (product, qty, date_obj, trans_type, rate, amount))
        conn.commit()

        inserted_id = cursor.lastrowid
        print("✅ Data inserted with ID:", inserted_id)

        # Treeview me insert — make sure your columns start with ID
        tree.insert("", "end", values=(inserted_id, product, qty, date_obj, trans_type, rate, amount))

        clear_fields()
        update_dashboard()

    except Exception as e:
        import traceback
        print("❌ Exception occurred:")
        traceback.print_exc()
        messagebox.showerror("Error", f"Error: {e}")

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def update_transaction():
    print("🟢 Step 1: update_transaction() called")
    selected = tree.focus()
    if not selected:
        messagebox.showwarning("No Selection", "⚠️ Please select a row to update.")
        return

    try:
        print("🟢 Step 2: Collecting input values")

        # ✅ Collect inputs
        product = product_cb.get().strip()
        qty_text = qty_entry.get().strip()
        rate_text = rate_entry.get().strip()
        trans_type = trans_cb.get().strip()

        # ✅ Get and format date properly for MySQL
        date = date_entry.get_date()
        date_str = date.strftime('%Y-%m-%d')  # MySQL-friendly format

        print("🔍 Collected:", product, qty_text, date_str, trans_type, rate_text)

        # ✅ Validate
        if not product or not qty_text or not rate_text or not trans_type:
            messagebox.showerror("Input Error", "❌ All fields must be filled before updating.")
            return

        qty = int(qty_text)
        rate = float(rate_text)
        amount = qty * rate

        # ✅ Get current values from Treeview
        values = tree.item(selected, 'values')
        transaction_id = values[0]

        # ✅ Compare new vs old to avoid unnecessary update
        current_values = (values[1], str(values[2]), str(values[3]), values[4], str(values[5]), str(values[6]))
        new_values = (product, str(qty), date_str, trans_type, f"{rate:.2f}", f"{amount:.2f}")

        if current_values == new_values:
            messagebox.showinfo("No Changes", "ℹ️ No changes detected to update.")
            return

        # ✅ Update in database
        conn = connect_db()
        cursor = conn.cursor()
        query = """
            UPDATE transactions 
            SET product=%s, qty=%s, date=%s, transaction_type=%s, rate=%s, amount=%s 
            WHERE id=%s
        """
        cursor.execute(query, (product, qty, date_str, trans_type, rate, amount, transaction_id))
        conn.commit()
        conn.close()

        messagebox.showinfo("Success", "✅ Transaction updated successfully.")
        refresh_data()
        clear_fields()

    except ValueError:
        messagebox.showerror("Conversion Error", "❌ Quantity and Rate must be valid numbers.")
    except Exception as e:
        messagebox.showerror("Update Error", f"❌ Error: {e}")



def delete_transaction():
    try:
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("No selection", "Please select a row to delete.")
            return

        confirm = messagebox.askyesno("Confirm", "Are you sure you want to delete this transaction?")
        if not confirm:
            return

        item = tree.item(selected)
        trans_id = item['values'][0]  # assuming 1st column is ID

        conn = connect_db()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM transactions WHERE id = %s", (trans_id,))
        conn.commit()
        conn.close()

        tree.delete(selected)
        update_dashboard()
        clear_fields()
        messagebox.showinfo("Deleted", "Transaction deleted successfully!")

    except Exception as e:
        messagebox.showerror("Delete Error", str(e))

def export_to_excel():
    messagebox.showinfo("Export", "Export to Excel is not yet implemented.")



def refresh_data():
    try:
        # Step 1: Clear Treeview
        tree.delete(*tree.get_children())

        # Step 2: Fetch fresh data from DB
        conn = connect_db()
        cursor = conn.cursor()

        cursor.execute("SELECT id, product, qty, date, transaction_type, rate, amount FROM transactions ORDER BY id DESC")
        rows = cursor.fetchall()

        for row in rows:
            tree.insert("", "end", values=(row[0], row[1], row[2], row[3], row[4], row[5], row[6]))

        conn.close()

        # Step 3: Update dashboard too
        update_dashboard()

        print("🔄 Data refreshed successfully.")

    except Exception as e:
        print("❌ Refresh Error:", e)
        messagebox.showerror("Error", f"Refresh Error:\n{e}")



def clear_fields():
    selected = tree.selection()
    
    if selected:
        confirm = messagebox.askyesno("Confirm Delete", "⚠️ Selected row delete karna chahte ho?")
        if confirm:
            item = tree.item(selected)
            values = item['values']
            trans_id = values[0]  # Assuming 1st column is ID

            try:
                conn = connect_db()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM transactions WHERE id = %s", (trans_id,))
                conn.commit()
                conn.close()

                tree.delete(selected)
                messagebox.showinfo("Deleted", "✅ Row deleted successfully.")
            except Exception as e:
                messagebox.showerror("Error", f"❌ DB Delete Error: {e}")

    # Clear form fields (always do this)
    product_cb.set("")
    qty_entry.delete(0, END)
    date_entry.set_date(datetime.now())
    trans_cb.set("")
    rate_entry.delete(0, END)
    tree.selection_remove(tree.selection())

    # 🟡 Optional: Check if all rows are deleted, then reset ID
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM transactions")
        count = cursor.fetchone()[0]
        conn.close()

        if count == 0:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("ALTER TABLE transactions AUTO_INCREMENT = 1")
            conn.commit()
            conn.close()
            print("✅ All rows deleted, ID reset to 1.")
    except Exception as e:
        print("⚠️ Auto-Increment Reset Error:", e)

# ================== Dashboard Frame ==================

dashboard = Frame(root, bg="skyblue")
dashboard.place(x=20, y=10, width=960, height=200)
# --- Report Frame ---
report_frame = LabelFrame(root, text="Report", bg="skyblue", font=("Arial", 10, "bold"), padx=20, pady=10)
report_frame.place(x=20, y=5, width=950, height=70)

def filter_report_by_date():
    try:
        conn = connect_db()
        cursor = conn.cursor()

        start = start_date.get_date().strftime('%Y-%m-%d')
        end = end_date.get_date().strftime('%Y-%m-%d')

        cursor.execute("""
            SELECT * FROM transactions
            WHERE date BETWEEN %s AND %s
        """, (start, end))

        rows = cursor.fetchall()

        tree.delete(*tree.get_children())
        for row in rows:
            tree.insert('', 'end', values=row)

        conn.close()
    except Exception as e:
        messagebox.showerror("Error", f"❌ Failed to filter report data\n{str(e)}")


from fpdf import FPDF

def generate_pdf():
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Inventory Report", ln=True, align="C")
        
        pdf.cell(200, 10, txt=f"From {start_date.get()} to {end_date.get()}", ln=True, align="C")
        pdf.ln(10)

        # Table Header
        headers = ["ID", "Product", "Qty", "Type", "Rate", "Amount", "Date"]
        col_widths = [10, 40, 20, 25, 25, 30, 35]

        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, border=1)

        pdf.ln()

        # Table Data
        for child in tree.get_children():
            values = tree.item(child)['values']
            for i, val in enumerate(values):
                pdf.cell(col_widths[i], 10, str(val), border=1)
            pdf.ln()

        pdf.output("Inventory_Report.pdf")
        messagebox.showinfo("PDF Saved", "📄 Report saved as Inventory_Report.pdf")
    except Exception as e:
        messagebox.showerror("Error", f"❌ PDF generation failed\n{str(e)}")



# ---------------- Report Frame ----------------
report_frame = LabelFrame(root, text="Report", bg="skyblue", font=("Arial", 10, "bold"), padx=20, pady=10)
report_frame.place(x=20, y=5, width=950, height=70)

Label(report_frame, text="Start Date:", bg="skyblue", font=("Arial", 10, "bold")).place(x=20, y=15)
start_date = DateEntry(report_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
start_date.place(x=100, y=15)

Label(report_frame, text="End Date:", bg="skyblue", font=("Arial", 10, "bold")).place(x=250, y=15)
end_date = DateEntry(report_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
end_date.place(x=320, y=15)

filter_btn = Button(report_frame, text="📅 Report Filter", font=("Arial", 10, "bold"), bg="#ffff00", command=filter_report_by_date)
filter_btn.place(x=420, y=10)

from PIL import Image, ImageTk
# Load PDF icon
pdf_icon_img = Image.open("pdf.png")
pdf_icon_img = pdf_icon_img.resize((20, 20))  # Resize if needed
pdf_icon = ImageTk.PhotoImage(pdf_icon_img)

# PDF Button with Icon
pdf_btn = Button(report_frame, image=pdf_icon, text=" PDF", compound=LEFT,
                 font=("Arial", 9, "bold"), bg="#0000ff", fg="white", command=generate_pdf)
pdf_btn.image = pdf_icon  # Prevent garbage collection
pdf_btn.place(x=565, y=10)


def update_dashboard():
    try:
        conn = connect_db()
        cursor = conn.cursor()

        cursor.execute("SELECT SUM(amount) FROM transactions WHERE transaction_type = 'Purchase'")
        purchase_total = cursor.fetchone()[0] or 0

        cursor.execute("SELECT SUM(amount) FROM transactions WHERE transaction_type = 'Sale'")
        sale_total = cursor.fetchone()[0] or 0

        profit = sale_total - purchase_total
        # Update the existing dashboard labels
        dashboard_labels[0].config(text=f"Purchase\n₹ {purchase_total:.2f}")
        dashboard_labels[1].config(text=f"Sale\n₹ {sale_total:.2f}")
        dashboard_labels[2].config(text=f"Profit\n₹ {profit:.2f}")
        # Inventory (Quantity and Value)
        cursor.execute("SELECT SUM(qty) FROM transactions WHERE transaction_type = 'Purchase'")
        purchase_qty = cursor.fetchone()[0] or 0
        cursor.execute("SELECT SUM(qty) FROM transactions WHERE transaction_type = 'Sale'")
        sale_qty = cursor.fetchone()[0] or 0
        
        inventory_qty = purchase_qty - sale_qty

        cursor.execute("SELECT SUM(qty * rate) FROM transactions WHERE transaction_type = 'Purchase'")
        purchase_amt = cursor.fetchone()[0] or 0
        cursor.execute("SELECT SUM(qty * rate) FROM transactions WHERE transaction_type = 'Sale'")
        sale_amt = cursor.fetchone()[0] or 0

        inventory_amt = purchase_amt - sale_amt

        dashboard_labels[3].config(text=f"Inventory Qty\n{inventory_qty}")
        dashboard_labels[4].config(text=f"Inventory Amt\n₹ {inventory_amt:.2f}")
        conn.close()  # ✅ close the connection


    except Exception as e:
        print("Dashboard update error:", e)

import openpyxl
from tkinter import filedialog

def export_to_excel():
    try:
        # Ask where to save file
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save as"
        )

        if not file_path:
            return  # Cancelled

        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Header
        headers = ["ID", "Product", "Qty", "Date", "Type", "Rate", "Amount"]
        ws.append(headers)

        # Fetch all rows from DB
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id, product, qty, date, transaction_type, rate, amount FROM transactions")
        rows = cursor.fetchall()
        conn.close()

        # Write to sheet
        for row in rows:
            ws.append(row)

        # Save
        wb.save(file_path)

        messagebox.showinfo("✅ Exported", f"Data exported successfully to:\n{file_path}")

    except Exception as e:
        messagebox.showerror("❌ Error", f"Failed to export:\n{e}")





dashboard_labels = []

dash_labels = [
    ("Purchase", 20, "green"),
    ("Sale", 190, "blue"),
    ("Profit", 360, "orange"),
    ("Inventory Qty", 530, "purple"),
    ("Inventory Amt", 700, "red")
]

for text, xpos, color in dash_labels:
    lbl = Label(dashboard, text=text, font=("Arial", 11, "bold"), bg=color, fg="white", relief="ridge", bd=2, justify=CENTER)
    lbl.place(x=xpos, y=80, width=150, height=50)
    dashboard_labels.append(lbl)

# update_dashboard()

# ================== Form Frame ==================
form_frame = LabelFrame(root, text="Sale / Purchase Entry", bg="skyblue", font=("Arial", 11, "bold"))
form_frame.place(x=20, y=160, width=960, height=140)

# ✅ Shadow Button Creator
def create_shadow_button(parent, text, x, y, bg, fg, command):
    btn_width = 110
    btn_height = 30
    offset = 2  # Shadow offset

    # Shadow (slightly offset behind the button)
    shadow = Canvas(parent, width=btn_width, height=btn_height, bg="white", highlightthickness=0)
    shadow.place(x=x + offset, y=y + offset)

    # Actual Button
    btn = Button(parent, text=text, width=12, font=("Arial", 11, "bold"),
                 bg=bg, fg=fg, bd=0, relief="flat", command=command)
    btn.place(x=x, y=y)

    return btn


# ✅ Create Buttons with Shadow
create_shadow_button(form_frame, "Add",    500, 45, "navy", "white", add_transaction)
create_shadow_button(form_frame, "Clear",  620, 45, "darkred", "white", clear_fields)
create_shadow_button(form_frame, "Update", 750, 45, "red", "white", update_transaction)


Label(form_frame, text="Product", bg="skyblue", font=("Arial", 10)).grid(row=0, column=0, padx=10, pady=10, sticky="w")
product_cb = ttk.Combobox(form_frame, values=["Pen", "Notebook", "Pencil"])
product_cb.grid(row=0, column=1, padx=10, pady=10)

Label(form_frame, text="Quantity", bg="skyblue", font=("Arial", 10)).grid(row=0, column=2, padx=10, pady=10, sticky="w")
qty_entry = Entry(form_frame)
qty_entry.grid(row=0, column=3, padx=10, pady=10)

Label(form_frame, text="Date", bg="skyblue", font=("Arial", 10)).grid(row=0, column=4, padx=10, pady=10, sticky="w")
date_entry = DateEntry(form_frame)
date_entry.grid(row=0, column=5, padx=10, pady=10)

Label(form_frame, text="Transaction Type", bg="skyblue", font=("Arial", 10)).grid(row=1, column=0, padx=10, pady=5, sticky="w")
trans_cb = ttk.Combobox(form_frame, values=["Sale", "Purchase"])
trans_cb.grid(row=1, column=1, padx=10, pady=5)

Label(form_frame, text="Rate", bg="skyblue", font=("Arial", 10)).grid(row=1, column=2, padx=10, pady=5, sticky="w")
rate_entry = Entry(form_frame)
rate_entry.grid(row=1, column=3, padx=10, pady=5)

Button(form_frame, text="Add", width=12, bg="navy", font=("Arial",11), fg="white", command=add_transaction).place(x=500, y=45)
Button(form_frame, text="Clear", width=12, bg="darkred", font=("Arial",11), fg="white", command=clear_fields).place(x=620, y=45)
Button(form_frame, text="Update", width=12, bg="red", fg="white", font=("Arial",11), command=update_transaction).place(x=750, y=45)

# ================== Filter Section ==================
filter_frame = LabelFrame(root, text="Search & Filter", bg="skyblue", font=("Arial", 11, "bold"))
filter_frame.place(x=20, y=270, width=960, height=100)

def apply_filter():
    try:
        start = start_date_entry.get_date().strftime('%Y-%m-%d')
        end = end_date_entry.get_date().strftime('%Y-%m-%d')
        field = select_cb.get()
        keyword = filter_entry.get().strip().lower()  # lowercase for consistency

        query = "SELECT * FROM transactions WHERE date BETWEEN %s AND %s"
        params = [start, end]

        if field == "Product":
            query += " AND product LIKE %s"
            params.append(f"%{keyword}%")
        elif field == "Transaction Type":
            query += " AND LOWER(transaction_type) = %s"
            params.append(keyword)  # already lowercased

        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()

        tree.delete(*tree.get_children())

        for row in rows:
            tree.insert('', 'end', values=row)

        conn.close()

    except Exception as e:
        messagebox.showerror("Filter Error", str(e))



def update_filter_input(event=None):
    selected = select_cb.get()

    # 🔄 Remove old filter input (row=0, column=7)
    for widget in filter_frame.grid_slaves(row=0, column=7):
        widget.destroy()

    # 🧠 Show proper input field based on selected filter
    global filter_entry
    if selected == "Transaction Type":
        filter_entry = ttk.Combobox(filter_frame, values=["purchase", "sale"])  # lowercase
        filter_entry.grid(row=0, column=7, padx=(2, 10), pady=10)
        filter_entry.set("purchase")  # default selection optional
    else:
        filter_entry = Entry(filter_frame)
        filter_entry.grid(row=0, column=7, padx=(2, 10), pady=10)


filter_entry = Entry(filter_frame)
filter_entry.grid(row=0, column=7, padx=(2, 10), pady=10)


filter_btn = Button(report_frame, text="Search Filter", font=("Arial", 10, "bold"),
                    bg="red", fg="white", command=apply_filter)
filter_btn.place(x=650, y=12)  # 👈 Yaha x/y manually adjust karo as per your layout


# Filter section inside filter_frame
Label(filter_frame, text="Start Date", bg="skyblue", font=("Arial", 10)).grid(row=0, column=0, padx=(10, 2), pady=10, sticky="e")
start_date_entry = DateEntry(filter_frame)
start_date_entry.grid(row=0, column=1, padx=(2, 10), pady=15)

Label(filter_frame, text="End Date", bg="skyblue", font=("Arial", 10)).grid(row=0, column=2, padx=(10, 2), pady=10, sticky="e")
end_date_entry = DateEntry(filter_frame)
end_date_entry.grid(row=0, column=3, padx=(2, 10), pady=10)

Label(filter_frame, text="Filter By", bg="skyblue", font=("Arial", 10)).grid(row=0, column=4, padx=(10, 2), pady=10, sticky="e")
select_cb = ttk.Combobox(filter_frame, values=["Product", "Transaction Type"])
select_cb.grid(row=0, column=5, padx=(2, 10), pady=10)
select_cb.set("Select")

Label(filter_frame, text="Keyword", bg="skyblue", font=("Arial", 10)).grid(row=0, column=6, padx=(10, 2), pady=10, sticky="e")
filter_entry = Entry(filter_frame)
filter_entry.grid(row=0, column=7, padx=(2, 10), pady=10)

try:
    icon_refresh = PhotoImage(file="refresh.png")
    icon_dustbin = PhotoImage(file="dustbin.png")
    icon_excel = PhotoImage(file="excel.png")  
    icon_edit = PhotoImage(file="edit.png")

    Button(filter_frame, width=25, height=25, image=icon_refresh, bg="white", command=refresh_data).place(x=785, y=10)
    Button(filter_frame, width=30, height=30, image=icon_dustbin, bg="white", command=delete_transaction).place(x=820, y=8)
    Button(filter_frame, image=icon_excel, width=30, height=30, bg="white", command=export_to_excel).place(x=865, y=8)
    Button(filter_frame, image=icon_edit, width=30, height=30, bg="white", command=update_transaction).place(x=910, y=8)
except Exception as e:
    print("Image Load Error:", e)


# ================== Table Frame ==================
table_frame = Frame(root, bg="white")
table_frame.place(x=20, y=360, width=960, height=280)

columns = ("ID", "Product", "Qty", "Date", "Transaction Type", "Rate", "Amount")
tree = ttk.Treeview(table_frame, columns=columns, show="headings")

for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=140, anchor="center")

from datetime import datetime

def on_row_select(event):
    selected = tree.selection()
    if selected:
        item = tree.item(selected)
        values = item['values']

        # Assuming your tree columns are: ID, Product, Qty, Date, Type, Rate, Amount
        product_cb.set(values[1])
        qty_entry.delete(0, END)
        qty_entry.insert(0, values[2])

        # ✅ Convert string to datetime.date before setting
        try:
            date_obj = datetime.strptime(values[3], "%Y-%m-%d").date()
            date_entry.set_date(date_obj)
        except ValueError as ve:
            print(f"❌ Date format error: {ve}")
            messagebox.showerror("Date Error", f"Invalid date format: {values[3]}")

        trans_cb.set(values[4])
        rate_entry.delete(0, END)
        rate_entry.insert(0, values[5])




tree.bind("<ButtonRelease-1>", on_row_select)



scroll_y = Scrollbar(table_frame, orient=VERTICAL, command=tree.yview)
tree.configure(yscrollcommand=scroll_y.set)
scroll_y.pack(side=RIGHT, fill=Y)
tree.pack(fill=BOTH, expand=True)

# ================== Run App ==================
def global_exception_handler(type, value, tb):
    import traceback
    err = ''.join(traceback.format_exception(type, value, tb))
    print("⚠️ Uncaught Exception:\n", err)
    with open("error_log.txt", "w") as f:
        f.write(err)
    messagebox.showerror("Critical Error", "Unexpected crash! Please check error_log.txt")

# import sys
# sys.excepthook = global_exception_handler


root.mainloop()


